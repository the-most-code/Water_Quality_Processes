# -*- coding: utf-8 -*-
"""
Created on Fri Feb  4 11:51:01 2022

@author: Tano_E
"""

'''
Multi-faceted python script with multiple uses.

1. TMDL Lake Modeling:
    Input: Watershed, rainfall, and folder location
    Output: Provides total level 2 landuse loading in excel files as well as bathtub file for TMDL development.
2. Nutrient Analysis:
    Input: Watershed, Waterbody, rainfall, avg number of people per household, and folder location
    Output: Provides total level 2 landuse loading shapefiles and level 1 landuse pie charts (includes septic). Level 2 landuse

'''
import os
import arcpy
import pandas as pd
import numpy as np
import xlsxwriter
from pathlib import Path
from collections import OrderedDict
import sys
from openpyxl import load_workbook

from math import pi
from bokeh.palettes import viridis
from bokeh.plotting import figure, show
from bokeh.transform import cumsum
from bokeh.io import save, output_file, export_png
from bokeh.layouts import column

arcpy.env.overwriteOutput = True

class PLSM:
    def __init__(self, watershed_input, rainfall_input, folder_location,
                 joinfile = pd.read_csv(r"\\fldep1\WQETP\TMDL\GIS_Tools\Statewide_landuse_masterlist_harper.csv"),
                 landuse_input = r"\\floridadep.net\GIS\GeoData\geopub\geopub.gdb\STATEWIDE_LANDUSE",
                 NHD_waterbody = r"\\floridadep.net\\GIS\\geodata\\geopub\\NHD.gdb\\Hydrography\\NHDWaterbody"):
        self.watershed = watershed_input
        self.rainfall = rainfall_input
        self.folder = folder_location
        self.joinfile = joinfile
        self.landuse = landuse_input
        self.NHD_waterbody = NHD_waterbody

    def rainfallQA(self):
        '''
        Takes rainfall csv input and performs QA to ensure columns are formatted properly.
        '''
        rainfall_df = pd.read_csv(self.rainfall)
        # update columns
        up_cl = ['Year' if 'year' in x.lower() else 'Total' for x in list(rainfall_df.columns)]
        # rename columns
        rainfall_df.columns = up_cl
        #rainfall_df.sortvalues('Year')
        col_names = ['Total', 'Year']

        for name in col_names:
            if name not in rainfall_df:
                arcpy.AddError(''''The column names in the rainfall table you provided do not exist.
                               You should have columns labeled: Total and Year. Please check your column names and try again.''')
                sys.exit()

        return rainfall_df

    def Clip(self, type = 'Model'):
        '''
        Takes watershed input and clips statewide landuse land cover to the watershed. The 'Analysis' dictates that intersecting waters
        between the NHDwaterbody24k layer and clipped layer will remove waters frmo analysis. Otherwise the 'model' input will output the
        landuse with water.
        '''
        # Create Annual File directory
        temp_folder_path = os.path.join(self.folder, 'PLSM_shapefiles')

        # Create folder if does not exist
        if not os.path.exists(temp_folder_path):
             os.mkdir(temp_folder_path)

        temp_out_clip = temp_folder_path + "\landuse_clip.shp"

        print('Clipping landuse to watershed')
        arcpy.Clip_analysis(self.landuse, self.watershed, temp_out_clip)

        if type == 'Analysis':
            intersect_waters = temp_folder_path + "\intersect.shp"
            arcpy.analysis.Intersect([temp_out_clip, self.NHD_waterbody], intersect_waters)

            layer_to_process = temp_folder_path + "\water_removed"
            arcpy.analysis.Erase(temp_out_clip, intersect_waters, layer_to_process)

        elif type == 'Model':
            layer_to_process = temp_out_clip

        return layer_to_process, temp_folder_path

    def Dissolve(self, layer_to_process, temp_folder_path):
        '''
        Takes clip output and dissolves by level 2 landuse codes to aggregate features.
        *Note that prior to clip function the statewide landuse layer has different column names than what is referenced in the 'dissolve_fields'
        variable. This is due to arcpro truncating the field names when arcpy operations are performed.
        '''
        ## Define parameters for dissolve
        clip_input = layer_to_process + ".shp"
        temp_out_dissolve = temp_folder_path + "\landuse_dissolve"

        dissolve_fields = ["LEVEL2_LAN", "LEVEL2_L_1"] #names are truncated to 10 characters when converted to fc in the initial clip

        print('Dissolving by level-2 landuse')
        arcpy.Dissolve_management(clip_input, temp_out_dissolve, dissolve_fields)
        ## Define parameters for addField and calculateField
        dissolve_input = temp_out_dissolve + ".shp"

        ## Get length of dissolve shapefile for later
        rows = [row for row in arcpy.da.SearchCursor(dissolve_input,'LEVEL2_LAN')]
        n_rows = len(rows)

        return dissolve_input, n_rows, clip_input

    def calculateField(self, dissolve_input):
        '''
        Takes dissolve output and calculates area in square meters for each level 2 landuse.
        '''
        ## Run AddField
        arcpy.AddField_management(dissolve_input, "Area_sq_m", "DOUBLE")

        print('Calculating landuse area')
        ## Run CalculateField
        exp = "!SHAPE.AREA@SQUAREMETERS!"
        arcpy.CalculateField_management(dissolve_input, "Area_sq_m", exp, "PYTHON")

    def attribute_to_CSV(self, dissolve_input, temp_folder_path):
        '''
        Follwing the calculateField() function the attribte table in arcpro is converted to a csv file.
        '''
        # Execute TableToTable
        arcpy.TableToTable_conversion(dissolve_input, temp_folder_path, "wshed_landuse.csv")

    def Merge(self, temp_folder_path, n_rows):
        '''
        Follwing the attribute_to_CSV() function the wshed csv is read and merged to class variable 'joinfile' to merge the area sq m associated
        with each landuse type to the rest of the Statewide_landuse_masterlist_harper.csv dataset.
        '''
        print("Merging tables")
        wshed_landuse = pd.read_csv(temp_folder_path + "\wshed_landuse.csv")

        join1 = pd.merge(wshed_landuse, self.joinfile, left_on = 'LEVEL2_LAN', right_on = 'LEVEL2_LANDUSE_CODE')
        ## Check to make sure all of the dissolved statewide landuse codes matched with ROC and EMC landuse codes. If not, send error message to terminate script.
        merged_rows = len(join1.LEVEL2_LAN)

        if n_rows != merged_rows:
            arcpy.AddError(''''Not all of the Statewide Landuse Codes in watershed matched with user-defined data.
                           Please make sure all land use codes and/or descriptions match and try again.''')
            sys.exit()
        ## Check to make sure there are no missing values in ROC and EMC columns
        missing_list = join1.columns[join1.isnull().any()].tolist()
        if not missing_list:
            pass
        else:
            arcpy.AddError("The following column(s) contain missing values: "+str(missing_list)+". Please check your data and try again.")
            sys.exit()

        return join1

    def writeData(self, rainfall_df, join1): # join1 for waterbody model, intersect_waters for nutrient analysis
        '''
        Takes the output from rainfallQA() and Merge() and writes results of calculations to the PLSM_raw.xlsx excel file.
        Within this function is nested function that will produce the PLSM_summary.xlsx file.
        '''
        arcpy.env.workspace = self.folder + r"\landuseLoading.gdb"

        filepath_raw = Path(self.folder + "\PLSM_raw.xlsx")
        if filepath_raw.is_file():
            arcpy.AddWarning(''''WARNING: PLSM raw file (in chosen location) already exists. Previous file was overwritten!
                             If you want to run this model for an additional watershed, please select another location.''')

        writer_raw = pd.ExcelWriter(self.folder + "\PLSM_raw.xlsx", engine='xlsxwriter')

        print("Calculating Nutrient Loads")
        ## Create dictionary to later extract values for each year provided
        dic = rainfall_df.set_index('Year').to_dict()['Total']

        ## Then order it
        ordered_dict = OrderedDict((k,dic.get(k)) for k in rainfall_df.Year)

        # Summary CSV - Has to be a better way of doing this
        Year = []
        Yearly_Runoff_Volume_m3 = []
        Yearly_TN_Load_kg = []
        Yearly_TP_Load_kg = []

        d = {}
        for k,v in ordered_dict.items():
            d[k] = pd.DataFrame([v])
            d[k].columns = ['Rainfall_in']
            d[k] = pd.concat([join1,d[k]],axis=1)
            d[k]['Rainfall_m'] = d[k]['Rainfall_in']*0.0254
            d[k]['Rainfall_Volume_m3'] = d[k]['Rainfall_m'][0]*d[k]['Area_sq_m']
            d[k]['Rainfall_Volume_L'] = d[k]['Rainfall_Volume_m3']*1000
            d[k]['Runoff_Volume_L'] = d[k]['Rainfall_Volume_L']*d[k]['ROC']
            d[k]['TN_Load_kg'] = d[k]['Runoff_Volume_L']*d[k]['EMC_TN']/1000000
            d[k]['TP_Load_kg'] = d[k]['Runoff_Volume_L']*d[k]['EMC_TP']/1000000

            d[k].to_excel(writer_raw, sheet_name = str(k))
            Year.append(str(k))
            Yearly_Runoff_Volume_m3.append(((d[k]['Runoff_Volume_L'])/1000).sum())
            Yearly_TN_Load_kg.append(d[k]['TN_Load_kg'].sum())
            Yearly_TP_Load_kg.append(d[k]['TP_Load_kg'].sum())

        writer_raw.save()
        def writeSummary(Year, Yearly_Runoff_Volume_m3, Yearly_TN_Load_kg, Yearly_TP_Load_kg):
            filepath_sum = Path(self.folder + "\PLSM_summary.xlsx")

            writer_sum = pd.ExcelWriter(self.folder + "\PLSM_summary.xlsx")

            if filepath_sum.is_file():
                arcpy.AddWarning('''WARNING: PLSM summary file (in chosen location) already exists. Previous file was overwritten!
                                 If you want to run this model for an additional watershed, please select another location.''')

            summary_df = pd.DataFrame({'Year': Year,
                                       'Yearly Runoff Volume (m^3)':Yearly_Runoff_Volume_m3,
                                       'Yearly TN Load (kg)':Yearly_TN_Load_kg,
                                       'Yearly TP Load (kg)':Yearly_TP_Load_kg})
            summary_df = summary_df.sort_values('Year')
            summary_df['Runoff Volume (hm3)'] = summary_df['Yearly Runoff Volume (m^3)']/1000000
            summary_df['Total T Conc (ppb)'] = summary_df['Yearly TN Load (kg)']/summary_df['Yearly Runoff Volume (m^3)']*1000000
            summary_df['Total P Conc (ppb)'] = summary_df['Yearly TP Load (kg)']/summary_df['Yearly Runoff Volume (m^3)']*1000000
            summary_df = summary_df.drop('Yearly TN Load (kg)', 1)
            summary_df = summary_df.drop('Yearly TP Load (kg)', 1)
            summary_df = summary_df.drop('Yearly Runoff Volume (m^3)', 1)

            n = len(summary_df.Year)+1
            data = summary_df[lambda summary_df: summary_df.columns[0:4]]

            pd.io.formats.excel.header_style = None

            data.to_excel(writer_sum, sheet_name='PLSM Summary', index=False)

            workbook  = writer_sum.book
            worksheet = writer_sum.sheets['PLSM Summary']

            font_fmt = workbook.add_format({'font_name': 'Calibri', 'font_size': 12, 'left':1, 'right':1, 'align': 'center', 'num_format': '0.00'})
            bottom_fmt = workbook.add_format({'font_name': 'Calibri', 'font_size': 12, 'bottom':1})
            border_fmt = workbook.add_format({'right': 1, 'left': 1})
            header_fmt = workbook.add_format({'font_name': 'Calibri Light', 'font_size': 14, 'bold': True, 'align':'center', 'bg_color':'#A4E1E2', 'bottom':1, 'right':1, 'left':1})
            bold = bold = workbook.add_format({'bold': True})
            superscript = workbook.add_format({'font_script':1,'bold': True})

            worksheet.set_default_row(hide_unused_rows=True)

            worksheet.set_column('A:A', 12, font_fmt)
            worksheet.set_column('B:B', 34, font_fmt)
            worksheet.set_column('C:D', 26, font_fmt)
            worksheet.set_column('E:XFD', None, None, {'hidden': True})


            for row in range(1, n):
                worksheet.set_row(row, 15, font_fmt)

            worksheet.write ('A1', 'Year', header_fmt)
            worksheet.write_rich_string ('B1', 'Runoff Volume (hm3)', header_fmt)

            worksheet.write ('C1', 'TN (ppb)', header_fmt)
            worksheet.write ('D1', 'TP (ppb)', header_fmt)

            writer_sum.save()
        writeSummary(Year, Yearly_Runoff_Volume_m3, Yearly_TN_Load_kg, Yearly_TP_Load_kg)

        return d

    def ltaLoading(self, dissolve_input, d):
        '''
        Takes output from Dissolve() and writeData() to produce a long-term average per acre representation of level 2 landuse loading.
        Exports shapefiles landuse_dissolveLTA_Loading to import into arcpro for spatial analysis.
        '''
        # Create excel file path
        writer_map = pd.ExcelWriter(self.folder + r"\LTA_LVL_2_Loading.xlsx", engine= 'xlsxwriter')
        # Initialize blank  dataframe
        lta_initial_df = pd.DataFrame()
        # Set columns for formatting
        lta = pd.DataFrame(columns= ['LEVEL2_LAN', 'LEVEL2_L_1', 'TN_Acre', 'TP_Acre'])
        for k in d.keys():
            # write loading per acre to lta dataframe
            lta[['LEVEL2_LAN', 'LEVEL2_L_1']] = d[k][['LEVEL2_LAN', 'LEVEL2_L_1']]
            # Convert sq^m to acree and calculate to get loading per acre of each landuse
            lta['TN_Acre'] = d[k]['TN_Load_kg'] / (d[k]['Area_sq_m']*0.00024711)
            lta['TP_Acre'] = d[k]['TP_Load_kg'] / (d[k]['Area_sq_m']*0.00024711)
            if len(lta_initial_df) < 1:
                # Copy first iteration of loading data to initialized dataframe
                lta_initial_df[['LEVEL2_LAN', 'LEVEL2_L_1', 'TN_Acre', 'TP_Acre']] = lta[['LEVEL2_LAN', 'LEVEL2_L_1', 'TN_Acre', 'TP_Acre']]
            else:
                # With data already written to initialized dataframe, sum the following loading data to it
                summed_columns = lta_initial_df[['TN_Acre', 'TP_Acre']] + lta[['TN_Acre', 'TP_Acre']]
                # Update initialized dataframe with summed data, further iteration will continue to sum data to perform long term average
                lta_initial_df.update(summed_columns, join= 'left', overwrite=True)
        # Get count of years for average
        yr_count = len(d.keys())
        # Divide loading data columns by year count to obtain average
        avg_loading = lta_initial_df[['TN_Acre', 'TP_Acre']] / yr_count
        # Have to update initialized df with average loading
        lta_initial_df.update(avg_loading, join= 'left', overwrite=True)
        # Setting index for formatting of arcpy table entry
        lta_initial_df = lta_initial_df.set_index('LEVEL2_LAN')
        # Drop waters from table
        lta_initial_df = lta_initial_df[(lta_initial_df[['TN_Acre', 'TP_Acre']] != 0).all(axis=1)]
        # Write the sheet to excel and save
        lta_initial_df.to_excel(writer_map, sheet_name = 'LTA Loading per Landuse')
        # Save excel file
        writer_map.save()

        lta_shp = arcpy.management.CopyFeatures(dissolve_input, dissolve_input[:-4] + 'LTA_Loading' + '.shp')
        join_table = arcpy.ExcelToTable_conversion(self.folder + r"\LTA_LVL_2_Loading.xlsx", self.folder + r"\landuseLoading" + ".gdb", 'LTA Loading per Landuse')
        join_field = ['TN_Acre', 'TP_Acre']
        # Joining the level 2 landuse data back to the shapefile creates 0 values in-place for water features that have no data
        arcpy.management.JoinField(lta_shp, 'LEVEL2_L_1', join_table, 'LEVEL2_L_1', join_field)
        # Delete water features by iterating through attribute table of feature
        with arcpy.da.UpdateCursor(lta_shp, ['TN_Acre', 'TP_Acre']) as cursor:
                for row in cursor:
                    if row[0] == 0:
                        cursor.deleteRow()
        # This is bugged and wont work in an arcpy script
        #symbology_layer = r'C:\development\Tool_development\SymbologyReference.lyrx'
        #arcpy.management.ApplySymbologyFromLayer(lta_shp, symbology_layer)

        return lta_initial_df

    def pieChart(self, d, clip_input, septic_loading = None, include_septic = False, remove_waters = False):
        '''
        Takes output from PLSM class writeData(), Dissolve(), and Septic class runCalculation() to produce a pie chart
        of long term average lvl 1 landuse. Function arguements determine wheter to include septic or water in loading representation.
        '''
        # long term average loading lvl 1 landuse
        writer_pie = pd.ExcelWriter(self.folder + r"\LVL_1_Landuse.xlsx", engine='xlsxwriter')
        # Initialize blank  dataframe
        lta_initial_df = pd.DataFrame()
        # Set columns for formatting
        lta = pd.DataFrame(columns= ['LEVEL2_LAN', 'LEVEL2_L_1', 'TN_Acre', 'TP_Acre'])
        for k in d.keys():
            # write loading per acre to lta dataframe
            lta[['LEVEL2_LAN', 'LEVEL2_L_1']] = d[k][['LEVEL2_LAN', 'LEVEL2_L_1']]
            # Convert sq^m to acree and calculate to get loading per acre of each landuse
            lta['TN_Acre'] = d[k]['TN_Load_kg']
            lta['TP_Acre'] = d[k]['TP_Load_kg']
            if len(lta_initial_df) < 1:
                # Copy first iteration of loading data to initialized dataframe
                lta_initial_df[['LEVEL2_LAN', 'LEVEL2_L_1', 'TN_Acre', 'TP_Acre']] = lta[['LEVEL2_LAN', 'LEVEL2_L_1', 'TN_Acre', 'TP_Acre']]
            else:
                # With data already written to initialized dataframe, sum the following loading data to it
                summed_columns = lta_initial_df[['TN_Acre', 'TP_Acre']] + lta[['TN_Acre', 'TP_Acre']]
                # Update initialized dataframe with summed data, further iteration will continue to sum data to perform long term average
                lta_initial_df.update(summed_columns, join= 'left', overwrite=True)
        # Get count of years for average
        yr_count = len(d.keys())
        #print(yr_count)
        # Divide loading data columns by year count to obtain average
        avg_loading = lta_initial_df[['TN_Acre', 'TP_Acre']] / yr_count
        # Have to update initialized df with average loading
        lta_initial_df.update(avg_loading, join= 'left', overwrite=True)
        # Setting index for formatting of arcpy table entry
        lta_initial_df = lta_initial_df.set_index('LEVEL2_LAN')

        def unique_values(table):
            with arcpy.da.SearchCursor(table, ['LEVEL1_LAN', 'LEVEL1_L_1', 'LEVEL2_LAN', 'LEVEL2_L_1']) as cursor:
                # with arcpy.da.SearchCursor(table, ['ESTUARY_SE']) as cursor: # i used this line to get the enr list csv to query my sqlite file
                return sorted({row for row in cursor})
        LU = unique_values(clip_input)

        LU_list = []
        for i in range(len(LU)):
            LU_list.append(LU[i])

        LU_df = pd.DataFrame(LU_list)

        landuse_df = LU_df.rename(columns={0:'LEVEL1_LAN', 1: 'LEVEL1_L_1',
                                           2: 'LEVEL2_LAN', 3:'LEVEL2_L_1'})

        lvl_LU_df = pd.merge(landuse_df, lta_initial_df, how= 'left', on=['LEVEL2_LAN', 'LEVEL2_L_1'])

        lvl_LU_df = lvl_LU_df.groupby(['LEVEL1_LAN', 'LEVEL1_L_1'])[['TN_Acre', 'TP_Acre']].agg(['mean']).reset_index()

        lvl_LU_df = lvl_LU_df.set_index('LEVEL1_LAN') # i hate working with pycharm for this reason
        lvl_LU_df = lvl_LU_df.droplevel(level=1, axis=1)

        lvl_LU_df = lvl_LU_df.rename(columns={'TN_Acre': 'TN_Kg',
                                              'TP_Acre': 'TP_Kg'})

        if include_septic == True:
            #lvl_LU_df.merge(septic_loading, left_on= 'LEVEL1_L_1', right_on= 'LEVEL1_L_1', how='inner') # this isnt merging the row
            dfs = [lvl_LU_df, septic_loading]
            # append septic loading to lvl 1 landuse for pie chart
            lvl_LU_df = pd.concat(dfs)

        if remove_waters == True:
            # Drop waters from table
            lvl_LU_df = lvl_LU_df[(lvl_LU_df[['TN_Kg', 'TP_Kg']] != 0).all(axis=1)]

        lvl_LU_df.to_excel(writer_pie, sheet_name='LVL 1 LTA Loading per Landuse')
        # Save excel file
        writer_pie.save()

        pies = []
        analytes = ['TN', 'TP']

        dir = self.folder + '\html_files'
        if not os.path.exists(dir):
            os.makedirs(dir)
        output_file(self.folder + "\\html_files\\Long Term Average Loading Kg.html")

        for a in analytes:
            data = lvl_LU_df[['LEVEL1_L_1', str(a) + '_Kg']]
            data['angle'] = data[str(a) + '_Kg']/data[str(a) + '_Kg'].sum() * 2*pi
            data['color'] = viridis(len(data))

            p = figure(height=950, width=1200, title= 'Long Term Average Loading ' + str(a) + ' Kg',
                   tools='hover', tooltips= '@LEVEL1_L_1: @'+ str(a) + '_Kg')

            p.wedge(x=0, y=1, radius=0.4,
                start_angle=cumsum('angle', include_zero=True), end_angle=cumsum('angle'),
                line_color='white', fill_color='color', legend_field='LEVEL1_L_1', source=data)

            p.axis.axis_label = None
            p.axis.visible = False
            p.legend.label_text_font_size = "8pt"
            p.grid.grid_line_color = None
            pies.append(p)
        plots = column(*pies)
        save(plots)
        #export_png(plots, dir)

    def annualLoading(self, dissolve_input, d):
        '''
        Takes output from Dissolve() and writeData() to produce an annual per acre representation of level 2 landuse loading.
        Exports shapefiles landuse_dissolve for every year there was rainfall data.
        '''
        # for testing purposes i can write this in a way that creates the excel file
        # when either this function or the ltaloading function is performed
        writer_map = pd.ExcelWriter(self.folder + r"\nutrientMap.xlsx", engine= 'xlsxwriter')

        for k in d.keys():
            nutrientMap = pd.DataFrame()
            nutrientMap[['LEVEL2_LAN', 'LEVEL2_L_1']] = d[k][['LEVEL2_LAN', 'LEVEL2_L_1']]
            nutrientMap['TN_Acre'] = d[k]['TN_Load_kg'] / (d[k]['Area_sq_m']*0.00024711) # converting area sq^m to acre
            nutrientMap['TP_Acre'] = d[k]['TP_Load_kg'] / (d[k]['Area_sq_m']*0.00024711)
            nutrientMap = nutrientMap.set_index('LEVEL2_LAN')
            nutrientMap.to_excel(writer_map, sheet_name = str(k))

        writer_map.save()
        # need to iterate over years and sheets of excel
        wb = load_workbook(self.folder + r"\nutrientMap.xlsx")


        #fieldObjList = arcpy.ListFields(self.watershed) # this loop here is good if the join is being applied to a table with a lot of unnecessary fields
       # fieldDelete = []

        #for field in fieldObjList:
        #    if not field.required:
        #        fieldDelete.append(field.name)

        for s in wb.worksheets:

            s = str(s)[12:-2]
            nutrientMap_shp = arcpy.management.CopyFeatures(dissolve_input, dissolve_input[:-4] + str(s) + '.shp')

            # keep this code in case of field deletion
            #print(fieldDelete)
            #arcpy.management.DeleteField(nutrientMap_shp, fieldDelete)

            join_table = arcpy.ExcelToTable_conversion(self.folder + r"\nutrientMap.xlsx", self.folder + r"\landuseLoading" + str(s) + ".gdb", str(s))
            join_field = ['TN_Acre', 'TP_Acre']

            arcpy.management.JoinField(nutrientMap_shp, 'LEVEL2_L_1', join_table, 'LEVEL2_L_1', join_field)

            # this file will have to be placed somewhere
            #symbologyFields = [["VALUE_FIELD", "TN_Acre", "TN_Acre"],
             #                  ["VALUE_FIELD", "TN_Acre", "TP_Acre"]]
            arcpy.management.ApplySymbologyFromLayer(nutrientMap_shp,
                                                     r'C:\development\Tool_development\SymbologyReference.lyrx')


            rename = 'nutrientLoading' + str(s) + '.shp'
            arcpy.management.Rename(nutrientMap_shp, rename)

    def plsm_data_extract(self):
        '''
        Produces bathtub csv from PLSM_raw and PLSM_summary excel files. This function doesnt need to have specified
        inputs for the excel files since they are being accessed from their file locations
        '''
        plsm_fl = self.folder +'\\PLSM_raw.xlsx'
        plsm_summ = self.folder +'\\PLSM_summary.xlsx'

        # Get sheetnames
        plsm_excel = pd.ExcelFile(plsm_fl)
        plsm_yr = plsm_excel.sheet_names

        # Get rainfall
        rain_value_list = []
        for yrs in plsm_yr:
            # Read the PLSM rawdata sheet
            plsm_sheets =  pd.read_excel(plsm_fl, sheet_name = yrs)
            # look up rainfall values
            rain_value = plsm_sheets['Rainfall_m'].iloc[0]
            # Append rain values to the list
            rain_value_list.append(rain_value)

        # Make a dataframe for rainfall only
        plsm_rainfall = pd.DataFrame({'Year':plsm_yr,'Precipitation  (meters)':rain_value_list})
        #plsm_rainfall.index = plsm_rainfall.Year

        # Get PLSM summary
        plsm_flows = pd.read_excel(plsm_summ)

        # Combine the dataframes
        plsm_join = pd.concat([plsm_rainfall,plsm_flows],axis = 1 ) #, how = 'left',left_on = 'Year' )

        # remove duplicate columns - Year
        plsm_join = plsm_join.loc[:,~plsm_join.columns.duplicated()]

        print("Getting PLSM results")

        # Get all PLSM data
        plsm_join.round(3).to_csv(self.folder + '\\PLSM_Bathtub.csv', index= False)

        return plsm_join



